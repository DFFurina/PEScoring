# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
import re
import os
import json
import warnings
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import math

warnings.filterwarnings("ignore")

CONFIG_FILE = "体育评分配置.json"

# ====================== 默认配置（还原所有可配置项） ======================
DEFAULT_CONFIG = {
    '姓名列': 'C', '性别列': 'D',
    '第一类项目列': 'E', '第一类成绩列': 'F', '第一类得分列': 'G',
    '第二类项目列': 'I', '第二类成绩列': 'J', '第二类得分列': 'K',
    '第三类项目列': 'M', '第三类成绩列': 'N', '第三类得分列': 'O',
    '第四类项目列': 'Q', '第四类成绩列': 'R', '第四类得分列': 'S',
    '总分列': 'T',
    '满分标准': 15.0,
    '优秀标准': 12.0,
    '是否加排名列': False,
    '缺考背景色': 'FFFF00',
    '立定跳远单位': '米',
    '实心球单位': '米',
    '自定义项目': [],
    '自定义映射项目': [],  # 新增：存储GUI添加的映射项目
    '总分小数位数': 2,
    '缺考字体颜色': 'FF0000',
    '排名并列方式': 'min',
    '是否自动排序班级': True,
    '统计包含优秀率': True,
    '统计文件名后缀': '_统计汇总',
    '当前版本': '2.0.0',
    '忽略更新版本': ''
}

STANDARD_FULL = {}  # 空字典，用于兼容GUI界面的循环调用

# ====================== 官方评分映射表（核心修复：替换线性计算为精准映射） ======================
# 1. 800/1000米（含附加分）- 官方标准（单位：秒）
RUN_SCORES = {
    '男生1000米': [
        (235, 6.0), (243, 5.7), (251, 5.4), (259, 5.1), (267, 4.8), (272, 4.5), (277, 4.2),
        (282, 3.9), (287, 3.6), (293, 3.3), (299, 3.0), (305, 2.7), (311, 2.4), (317, 2.1),
        (323, 1.8), (329, 1.5), (335, 1.2), (341, 0.9), (347, 0.6), (353, 0.3), (359, 0.0)
    ],
    '女生800米': [
        (220, 6.0), (228, 5.7), (236, 5.4), (244, 5.1), (252, 4.8), (257, 4.5), (262, 4.2),
        (267, 3.9), (272, 3.6), (278, 3.3), (284, 3.0), (290, 2.7), (296, 2.4), (302, 2.1),
        (308, 1.8), (314, 1.5), (320, 1.2), (326, 0.9), (332, 0.6), (338, 0.3), (344, 0.0)
    ]
}

RUN_BONUS = {
    '女生800米_1分秒数': 205,    # 3'25"及以内加1分
    '女生800米_0.5分秒数': 219,  # 3'26"-3'39"加0.5分
    '男生1000米_1分秒数': 220,  # 3'40"及以内加1分
    '男生1000米_0.5分秒数': 234  # 3'41"-3'54"加0.5分
}

# 2. 实心球（双手头上前掷）- 官方标准（单位：米）
SHIXINQIU_SCORES = {
    '男生': [
        (9.7, 3.0), (9.4, 2.85), (9.1, 2.7), (8.8, 2.55), (8.5, 2.4), (8.2, 2.25), (7.9, 2.1),
        (7.6, 1.95), (7.3, 1.8), (7.0, 1.65), (6.7, 1.5), (6.4, 1.35), (6.1, 1.2), (5.8, 1.05),
        (5.5, 0.9), (5.2, 0.75), (4.9, 0.6), (4.6, 0.45), (4.3, 0.3), (4.0, 0.15)
    ],
    '女生': [
        (6.8, 3.0), (6.6, 2.85), (6.4, 2.7), (6.2, 2.55), (6.0, 2.4), (5.8, 2.25), (5.6, 2.1),
        (5.4, 1.95), (5.2, 1.8), (5.0, 1.65), (4.8, 1.5), (4.6, 1.35), (4.4, 1.2), (4.2, 1.05),
        (4.0, 0.9), (3.8, 0.75), (3.6, 0.6), (3.4, 0.45), (3.2, 0.3), (3.0, 0.15)
    ]
}

# 3. 立定跳远 - 官方标准（单位：米）
LIDINGTIAOYUAN_SCORES = {
    '男生': [
        (2.49, 3.0), (2.41, 2.85), (2.33, 2.7), (2.25, 2.55), (2.20, 2.4), (2.15, 2.25), (2.10, 2.1),
        (2.06, 1.95), (2.02, 1.8), (1.98, 1.65), (1.94, 1.5), (1.90, 1.35), (1.86, 1.2), (1.82, 1.05),
        (1.78, 0.9), (1.74, 0.75), (1.70, 0.6), (1.66, 0.45), (1.62, 0.3), (1.58, 0.15)
    ],
    '女生': [
        (1.99, 3.0), (1.93, 2.85), (1.87, 2.7), (1.81, 2.55), (1.77, 2.4), (1.73, 2.25), (1.69, 2.1),
        (1.65, 1.95), (1.61, 1.8), (1.57, 1.65), (1.53, 1.5), (1.49, 1.35), (1.45, 1.2), (1.41, 1.05),
        (1.37, 0.9), (1.33, 0.75), (1.28, 0.6), (1.23, 0.45), (1.18, 0.3), (1.13, 0.15)
    ]
}

# 4. 50米跑 - 官方标准（单位：秒），时间越小得分越高
WUMISHI_SCORES = {
    '男生': [
        (7.1, 3.0), (7.2, 2.85), (7.3, 2.7), (7.4, 2.55), (7.5, 2.4), (7.6, 2.25), (7.7, 2.1),
        (7.9, 1.95), (8.1, 1.8), (8.3, 1.65), (8.5, 1.5), (8.7, 1.35), (8.9, 1.2), (9.1, 1.05),
        (9.3, 0.9), (9.5, 0.75), (9.7, 0.6), (9.9, 0.45), (10.1, 0.3), (10.3, 0.15)
    ],
    '女生': [
        (8.1, 3.0), (8.2, 2.85), (8.3, 2.7), (8.4, 2.55), (8.5, 2.4), (8.6, 2.25), (8.8, 2.1),
        (9.0, 1.95), (9.2, 1.8), (9.4, 1.65), (9.6, 1.5), (9.8, 1.35), (10.0, 1.2), (10.2, 1.05),
        (10.4, 0.9), (10.6, 0.75), (10.8, 0.6), (11.0, 0.45), (11.2, 0.3), (11.4, 0.15)
    ]
}

# 5. 25米游泳 - 官方标准（单位：秒）
YOUYONG25_SCORES = {
    '男生': [
        (22.0, 3.0), (23.0, 2.85), (24.0, 2.7), (25.0, 2.55), (26.0, 2.4), (27.0, 2.25), (28.0, 2.1),
        (29.0, 1.95), (30.0, 1.8), (31.0, 1.65), (32.0, 1.5), (33.0, 1.35), (34.0, 1.2), (35.0, 1.05),
        (36.0, 0.9), (37.0, 0.75), (38.0, 0.6), (39.0, 0.45), (40.0, 0.3), (41.0, 0.15)
    ],
    '女生': [
        (25.0, 3.0), (26.0, 2.85), (27.0, 2.7), (28.0, 2.55), (29.0, 2.4), (30.0, 2.25), (31.0, 2.1),
        (32.0, 1.95), (33.0, 1.8), (34.0, 1.65), (35.0, 1.5), (36.0, 1.35), (37.0, 1.2), (38.0, 1.05),
        (39.0, 0.9), (40.0, 0.75), (41.0, 0.6), (42.0, 0.45), (43.0, 0.3), (44.0, 0.15)
    ]
}

# 6. 200米游泳 - 官方标准（单位：秒）
YOUYONG200_SCORES = {
    '男生': [
        (276, 6.0), (288, 5.7), (300, 5.4), (312, 5.1), (324, 4.8), (336, 4.5), (348, 4.2),
        (360, 3.9), (372, 3.6), (384, 3.3), (396, 3.0), (408, 2.7), (420, 2.4), (432, 2.1),
        (444, 1.8), (456, 1.5), (468, 1.2), (480, 0.9), (492, 0.6), (504, 0.3), (516, 0.0)
    ],
    '女生': [
        (296, 6.0), (308, 5.7), (320, 5.4), (332, 5.1), (344, 4.8), (356, 4.5), (368, 4.2),
        (380, 3.9), (392, 3.6), (404, 3.3), (416, 3.0), (428, 2.7), (440, 2.4), (452, 2.1),
        (464, 1.8), (476, 1.5), (488, 1.2), (500, 0.9), (512, 0.6), (524, 0.3), (536, 0.0)
    ]
}

# 7. 足球运球 - 官方标准（单位：秒）
ZUQIU_SCORES = {
    '男生': [
        (7.6, 3.0), (8.7, 2.85), (9.6, 2.7), (10.5, 2.55), (11.3, 2.4), (12.2, 2.25), (13.1, 2.1),
        (14.3, 1.95), (15.5, 1.8), (16.0, 1.65), (16.8, 1.5), (17.4, 1.35), (17.9, 1.2), (18.3, 1.05),
        (19.0, 0.9), (19.6, 0.75), (20.0, 0.6), (20.5, 0.45), (21.0, 0.3), (21.5, 0.15)
    ],
    '女生': [
        (8.5, 3.0), (10.8, 2.85), (12.9, 2.7), (14.2, 2.55), (16.4, 2.4), (18.3, 2.25), (19.5, 2.1),
        (20.7, 1.95), (22.0, 1.8), (22.8, 1.65), (23.5, 1.5), (23.9, 1.35), (24.6, 1.2), (25.1, 1.05),
        (25.5, 0.9), (26.2, 0.75), (26.8, 0.6), (27.3, 0.45), (27.9, 0.3), (28.3, 0.15)
    ]
}

# 8. 篮球 - 官方标准（单位：秒）
LANQIU_SCORES = {
    '男生': [
        (20, 3.0), (21, 2.85), (22, 2.7), (23, 2.55), (24, 2.4), (25, 2.25), (27, 2.1),
        (29, 1.95), (31, 1.8), (33, 1.65), (35, 1.5), (37, 1.35), (39, 1.2), (41, 1.05),
        (43, 0.9), (45, 0.75), (47, 0.6), (49, 0.45), (51, 0.3), (53, 0.15)
    ],
    '女生': [
        (26, 3.0), (27, 2.85), (28, 2.7), (30, 2.55), (32, 2.4), (34, 2.25), (36, 2.1),
        (38, 1.95), (42, 1.8), (46, 1.65), (50, 1.5), (54, 1.35), (58, 1.2), (62, 1.05),
        (66, 0.9), (70, 0.75), (74, 0.6), (78, 0.45), (82, 0.3), (86, 0.15)
    ]
}

# 9. 排球垫球（40秒）- 官方标准（单位：次）
PAIQIU_SCORES = [
    (45, 3.0), (43, 2.85), (40, 2.7), (37, 2.55), (34, 2.4), (31, 2.25), (29, 2.1),
    (26, 1.95), (23, 1.8), (20, 1.65), (18, 1.5), (16, 1.35), (14, 1.2), (12, 1.05),
    (10, 0.9), (8, 0.75), (6, 0.6), (5, 0.45), (4, 0.3), (3, 0.15)
]

# 10. 乒乓球 - 官方标准（单位：次）
PINGPONG_SCORES = {
    30: 3.0, 29: 3.0, 28: 3.0, 27: 3.0, 26: 3.0, 25: 3.0,
    24: 2.88, 23: 2.76, 22: 2.64, 21: 2.52, 20: 2.40,
    19: 2.28, 18: 2.16, 17: 2.04, 16: 1.92, 15: 1.80,
    14: 1.68, 13: 1.56, 12: 1.44, 11: 1.32, 10: 1.20,
    9: 1.08, 8: 0.96, 7: 0.84, 6: 0.72, 5: 0.60,
    4: 0.48, 3: 0.36, 2: 0.24, 1: 0.12, 0: 0.0
}

# 11. 羽毛球 - 官方标准（单位：分）
YUMAOQIU_SCORES = [
    (150, 3.0), (149, 2.95), (140, 2.9), (130, 2.85), (120, 2.8), (110, 2.75),
    (100, 2.7), (95, 2.6), (90, 2.5), (85, 2.4), (80, 2.3), (75, 2.2),
    (70, 2.1), (65, 2.0), (60, 1.9), (55, 1.8), (50, 1.7), (45, 1.6),
    (40, 1.5), (35, 1.4), (30, 1.3), (25, 1.2), (20, 1.1), (15, 1.0),
    (10, 0.9), (5, 0.8), (0, 0.0)
]

# 12. 网球 - 官方标准（单位：分）
WANGQIU_SCORES = {
    '男生': [
        (120, 3.0), (99, 2.7), (88, 2.4), (77, 2.1), (66, 1.8), (55, 1.5),
        (44, 1.2), (33, 0.9), (22, 0.6), (11, 0.3), (8, 0.15), (0, 0.0)
    ],
    '女生': [
        (120, 3.0), (91, 2.7), (81, 2.4), (71, 2.1), (61, 1.8), (51, 1.5),
        (41, 1.2), (31, 0.9), (21, 0.6), (11, 0.3), (8, 0.15), (0, 0.0)
    ]
}

# 13. 仰卧起坐（1分钟）- 官方标准（单位：次）
YANGWOQIZUO_SCORES = [
    (50, 3.0), (47, 2.85), (44, 2.7), (41, 2.55), (38, 2.4), (35, 2.25),
    (32, 2.1), (30, 1.95), (28, 1.8), (26, 1.65), (24, 1.5), (22, 1.35),
    (20, 1.2), (18, 1.05), (16, 0.9), (14, 0.75), (12, 0.6), (10, 0.45),
    (8, 0.3), (6, 0.15)
]

# 14. 引体向上 - 官方标准（单位：次）
YINTIXIANGSHANG_SCORES = [
    (11, 3.0), (10, 2.85), (9, 2.7), (8, 2.4), (7, 2.1), (6, 1.8),
    (5, 1.5), (4, 1.2), (3, 0.9), (2, 0.6), (1, 0.3)
]

# 15. 4分钟跳绳 - 官方标准（单位：次）
TIAOSHENG_SCORES = {
    '男生': [
        (400, 6.0), (395, 5.7), (390, 5.4), (385, 5.1), (380, 4.8), (370, 4.5),
        (365, 4.2), (360, 3.9), (340, 3.6), (320, 3.3), (300, 3.0), (290, 2.7),
        (285, 2.4), (280, 2.1), (275, 1.8), (270, 1.5), (265, 1.2), (260, 0.9),
        (255, 0.6), (250, 0.3)
    ],
    '女生': [
        (405, 6.0), (400, 5.7), (395, 5.4), (390, 5.1), (385, 4.8), (375, 4.5),
        (370, 4.2), (365, 3.9), (345, 3.6), (325, 3.3), (305, 3.0), (295, 2.7),
        (290, 2.4), (285, 2.1), (280, 1.8), (275, 1.5), (270, 1.2), (265, 0.9),
        (260, 0.6), (250, 0.3)
    ]
}

# 16. 武术/体操 - 官方标准（单位：分，满分10分换算为3分制）
WUSHU_TICAO_SCORES = [
    (9.0, 3.0), (7.5, 2.25), (6.0, 1.8), (5.9, 0.0)
]

HEADERS = ['序号', '班级', '姓名', '性别',
           '第一类项目', '第一类成绩', '第一类得分', '',
           '第二类项目', '第二类成绩', '第二类得分', '',
           '第三类项目', '第三类成绩', '第三类得分', '',
           '第四类项目', '第四类成绩', '第四类得分', '总分']

# ====================== 工具函数 ======================
def col_to_num(col):
    num = 0
    for c in col.upper():
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num - 1

def time_to_seconds(s):
    if pd.isna(s) or str(s).strip() == '': return np.nan
    s = str(s).strip().replace('′', "'").replace('″', '"').replace(':', "'").replace('’', "'")
    # 新增：处理纯数字（如“316”→“3'16"”，“408”→“4'08"”）
    if s.isdigit() and len(s) >=3:
        m = s[:-2]  # 分钟部分
        sec = s[-2:] # 秒数部分
        s = f"{m}'{sec}"
    s = re.sub(r'[^\d\'"]', '', s)
    try:
        if "'" in s:
            m, rest = s.split("'")
            sec = float(rest.replace('"', ''))
            return int(m) * 60 + sec
        return float(s)
    except:
        return np.nan

def to_num(x):
    if pd.isna(x): return np.nan
    return float(re.sub(r'[^\d.]', '', str(x)))

# ====================== 核心评分函数（重构：全部使用官方映射表） ======================
def get_score(gender, project, value):
    if pd.isna(value) or str(value).strip() in ['', '-']:
        return "缺考"

    p = str(project).strip()
    # 项目名称标准化（兼容不同命名）
    p_map = {
        '800米': '800米', '1000米': '1000米', '800/1000米': '800米' if gender == '女' else '1000米',
        '实心球': '实心球', '双手头上前掷实心球': '实心球',
        '仰卧起坐': '仰卧起坐', '一分钟仰卧起坐': '仰卧起坐',
        '4分钟跳绳': '4分钟跳绳', '4min跳绳': '4分钟跳绳',
        '50米跑': '50米跑', '50m跑': '50米跑',  # 兼容“50m跑”
        '排球垫球': '排球', '排球': '排球',
        '乒乓': '乒乓球', '乒乓球': '乒乓球',
        '25m游泳': '25米游泳', '25米游泳': '25米游泳',
        '200m游泳': '200米游泳', '200米游泳': '200米游泳',
        '武术操': '武术', '武术': '武术', '体操': '体操', '羽毛球': '羽毛球',
        '篮球': '篮球', '足球': '足球运球', '足球运球': '足球运球',
        '网球': '网球', '引体向上': '引体向上', '立定跳远': '立定跳远'
    }
    p = p_map.get(p, p)

    # 转换成绩为数值
    val = to_num(value)
    if np.isnan(val):
        return "缺考"

    # 单位转换
    if p == '立定跳远' and DEFAULT_CONFIG['立定跳远单位'] == '厘米':
        val /= 100
    if p == '实心球' and DEFAULT_CONFIG['实心球单位'] == '厘米':
        val /= 100

    # 优先匹配GUI添加的自定义映射项目
    for custom in DEFAULT_CONFIG['自定义映射项目']:
        if custom['name'] == p:
            # 按映射表查找得分
            for lim, sc in custom['score_map']:
                if (custom['bigger_better'] and val >= lim) or (not custom['bigger_better'] and val <= lim):
                    return round(sc, 2)
            return 0.0

    # 原自定义项目（比例制）兼容
    for custom in DEFAULT_CONFIG['自定义项目']:
        if custom['name'] == p:
            std = custom['full_value']
            if std == 0: return 0.0
            ratio = val / std if custom['bigger_better'] else (std / val if val > 0 else 0)
            score = ratio * 3
            return round(max(0, min(3, score)), 2)

    # ====================== 恢复附加分逻辑 ======================
    # 800米/1000米（基础分+附加分）
    if p in ['800米', '1000米']:
        sec = time_to_seconds(value)
        if np.isnan(sec): return "缺考"
        
        # 基础分
        base_score = 0.0
        if p == '1000米' and gender == '男':
            for lim, sc in RUN_SCORES['男生1000米']:
                if sec <= lim:
                    base_score = sc
                    break
        elif p == '800米' and gender == '女':
            for lim, sc in RUN_SCORES['女生800米']:
                if sec <= lim:
                    base_score = sc
                    break
        
        # 恢复原附加分判断逻辑
        bonus = 0.0
        if p == '1000米' and gender == '男':
            bonus = 1.0 if sec <= RUN_BONUS['男生1000米_1分秒数'] else (0.5 if sec <= RUN_BONUS['男生1000米_0.5分秒数'] else 0.0)
        elif p == '800米' and gender == '女':
            bonus = 1.0 if sec <= RUN_BONUS['女生800米_1分秒数'] else (0.5 if sec <= RUN_BONUS['女生800米_0.5分秒数'] else 0.0)
        
        total = min(base_score + bonus, 7.0)
        return round(total, 2)

    # 2. 实心球（第二类项目，满分3分）
    if p == '实心球':
        scores = SHIXINQIU_SCORES['男生'] if gender == '男' else SHIXINQIU_SCORES['女生']
        for lim, sc in scores:
            if val >= lim:
                return round(sc, 2)
        return 0.0

    # 3. 立定跳远（第二类项目，满分3分）
    if p == '立定跳远':
        scores = LIDINGTIAOYUAN_SCORES['男生'] if gender == '男' else LIDINGTIAOYUAN_SCORES['女生']
        for lim, sc in scores:
            if val >= lim:
                return round(sc, 2)
        return 0.0

    # 4. 50米跑（第二类项目，满分3分）
    if p == '50米跑':
        scores = WUMISHI_SCORES['男生'] if gender == '男' else WUMISHI_SCORES['女生']
        for lim, sc in scores:
            if val <= lim:  # 时间越小得分越高
                return round(sc, 2)
        return 0.0

    # 5. 25米游泳（第二类项目，满分3分）
    if p == '25米游泳':
        sec = time_to_seconds(value)
        if np.isnan(sec): return "缺考"
        scores = YOUYONG25_SCORES['男生'] if gender == '男' else YOUYONG25_SCORES['女生']
        for lim, sc in scores:
            if sec <= lim:
                return round(sc, 2)
        return 0.0

    # 6. 引体向上（男生第二类项目，满分3分）
    if p == '引体向上' and gender == '男':
        for lim, sc in YINTIXIANGSHANG_SCORES:
            if val >= lim:
                return round(sc, 2)
        return 0.0

    # 7. 仰卧起坐（女生第二类项目，满分3分）
    if p == '仰卧起坐' and gender == '女':
        for lim, sc in YANGWOQIZUO_SCORES:
            if val >= lim:
                return round(sc, 2)
        return 0.0

    # 8. 4分钟跳绳（第一类项目，满分6分）
    if p == '4分钟跳绳':
        scores = TIAOSHENG_SCORES['男生'] if gender == '男' else TIAOSHENG_SCORES['女生']
        for lim, sc in scores:
            if val >= lim:
                return round(sc, 2)
        return 0.0

    # 9. 200米游泳（第一类项目，满分6分）
    if p == '200米游泳':
        sec = time_to_seconds(value)
        if np.isnan(sec): return "缺考"
        scores = YOUYONG200_SCORES['男生'] if gender == '男' else YOUYONG200_SCORES['女生']
        for lim, sc in scores:
            if sec <= lim:
                return round(sc, 2)
        return 0.0

    # 10. 乒乓球（第三类项目，满分3分）
    if p == '乒乓球':
        n = int(val) if val >= 0 else 0
        return PINGPONG_SCORES.get(n, 0.0)

    # 11. 羽毛球（第三类项目，满分3分）
    if p == '羽毛球':
        for lim, sc in YUMAOQIU_SCORES:
            if val >= lim:
                return round(sc, 2)
        return 0.0

    # 12. 网球（第三类项目，满分3分）
    if p == '网球':
        scores = WANGQIU_SCORES['男生'] if gender == '男' else WANGQIU_SCORES['女生']
        for lim, sc in scores:
            if val >= lim:
                return round(sc, 2)
        return 0.0

    # 13. 武术/体操（第三类项目，满分3分）
    if p in ['武术', '体操']:
        for lim, sc in WUSHU_TICAO_SCORES:
            if val >= lim:
                return round(sc, 2)
        return 0.0

    # 14. 足球运球（第四类项目，满分3分）
    if p == '足球运球':
        sec = time_to_seconds(value)
        if np.isnan(sec): return "缺考"
        scores = ZUQIU_SCORES['男生'] if gender == '男' else ZUQIU_SCORES['女生']
        for lim, sc in scores:
            if sec <= lim:
                return round(sc, 2)
        return 0.0

    # 15. 篮球（第四类项目，满分3分）
    if p == '篮球':
        sec = time_to_seconds(value)
        if np.isnan(sec): return "缺考"
        scores = LANQIU_SCORES['男生'] if gender == '男' else LANQIU_SCORES['女生']
        for lim, sc in scores:
            if sec <= lim:
                return round(sc, 2)
        return 0.0

    # 16. 排球（第四类项目，满分3分）
    if p == '排球':
        for lim, sc in PAIQIU_SCORES:
            if val >= lim:
                return round(sc, 2)
        return 0.0

    # 未匹配到的项目（默认返回0分，避免报错）
    print(f"未匹配到项目：{gender} - {p} - {value}")
    return 0.0

# ====================== GUI 类（新增更多配置项） ======================
class SportsScoreGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("上海中考体育评分")
        self.root.geometry("1000x800")

        self.input_file = ""
        self.output_file = ""
        self.stats_file = ""
        self.sheets = []
        self.sheet_vars = []
        self.config_entries = {}
        self.standard_entries = {}
        self.custom_projects = []

        self.var_statistics = tk.IntVar()

        self.create_widgets()
        self.load_config()
        # self.update_custom_list()
        self.check_for_update()
        # 初始化新版列表（替换旧的 update_custom_list）
        self.update_map_list()    # 初始化映射式项目列表
        self.update_ratio_list()  # 初始化比例式项目列表


    def create_widgets(self):
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)

        # 选项卡1: 主要操作
        tab_main = ttk.Frame(notebook)
        notebook.add(tab_main, text="主要操作")

        ttk.Label(tab_main, text="输入文件:", font=("Arial", 12)).grid(row=0, column=0, padx=10, pady=10, sticky="e")
        self.input_entry = ttk.Entry(tab_main, width=70)
        self.input_entry.grid(row=0, column=1, padx=10, pady=10)
        ttk.Button(tab_main, text="浏览...", command=self.select_input).grid(row=0, column=2, padx=10, pady=10)

        ttk.Label(tab_main, text="选择班级:", font=("Arial", 12)).grid(row=1, column=0, padx=10, pady=10, sticky="ne")
        self.sheet_frame = ttk.Frame(tab_main)
        self.sheet_frame.grid(row=1, column=1, columnspan=2, padx=10, pady=10, sticky="w")

        ttk.Label(tab_main, text="输出文件:", font=("Arial", 12)).grid(row=2, column=0, padx=10, pady=10, sticky="e")
        self.output_entry = ttk.Entry(tab_main, width=70)
        self.output_entry.grid(row=2, column=1, padx=10, pady=10)
        ttk.Button(tab_main, text="浏览...", command=self.select_output).grid(row=2, column=2, padx=10, pady=10)

        stats_frame = ttk.LabelFrame(tab_main, text="统计汇总")
        stats_frame.grid(row=3, column=0, columnspan=3, padx=10, pady=10, sticky="ew")
        ttk.Checkbutton(stats_frame, text="生成班级统计汇总（单独文件）", variable=self.var_statistics, command=self.toggle_stats_path).pack(anchor="w", padx=10, pady=5)
        self.stats_entry = ttk.Entry(stats_frame, width=70)
        self.stats_entry.pack(padx=10, pady=5, fill="x")
        self.stats_entry.config(state="disabled")
        ttk.Button(stats_frame, text="浏览...", command=self.select_stats_output).pack(pady=5)

        ttk.Button(tab_main, text="开始评分", command=self.process, style="Accent.TButton").grid(row=4, column=1, pady=30)
        ttk.Button(tab_main, text="保存当前配置", command=self.save_config).grid(row=5, column=1, pady=10)

        # 选项卡2: 列配置
        tab_config = ttk.Frame(notebook)
        notebook.add(tab_config, text="列配置")

        canvas_config = tk.Canvas(tab_config)
        scrollbar_config = ttk.Scrollbar(tab_config, orient="vertical", command=canvas_config.yview)
        scrollable_frame_config = ttk.Frame(canvas_config)

        scrollable_frame_config.bind("<Configure>", lambda e: canvas_config.configure(scrollregion=canvas_config.bbox("all")))

        canvas_config.create_window((0, 0), window=scrollable_frame_config, anchor="nw")
        canvas_config.configure(yscrollcommand=scrollbar_config.set)

        config_frame = ttk.LabelFrame(scrollable_frame_config, text="自定义列字母（大写字母，如 C、T）")
        config_frame.pack(fill="both", expand=True, padx=20, pady=20)

        row = 0
        for label, default in DEFAULT_CONFIG.items():
            if label in ['满分标准', '优秀标准', '是否加排名列', '缺考背景色', '立定跳远单位', '实心球单位', '自定义项目']:
                continue
            ttk.Label(config_frame, text=label + ":").grid(row=row, column=0, padx=10, pady=5, sticky="e")
            entry = ttk.Entry(config_frame, width=10)
            entry.insert(0, str(default))
            entry.grid(row=row, column=1, padx=10, pady=5)
            self.config_entries[label] = entry
            row += 1

        canvas_config.pack(side="left", fill="both", expand=True)
        scrollbar_config.pack(side="right", fill="y")

        # 评分标准配置选项卡
        tab_standard = ttk.Frame(notebook)
        notebook.add(tab_standard, text="评分标准配置")

        canvas_standard = tk.Canvas(tab_standard)
        scrollbar_standard = ttk.Scrollbar(tab_standard, orient="vertical", command=canvas_standard.yview)
        scrollable_frame_standard = ttk.Frame(canvas_standard)

        scrollable_frame_standard.bind("<Configure>", lambda e: canvas_standard.configure(scrollregion=canvas_standard.bbox("all")))

        canvas_standard.create_window((0, 0), window=scrollable_frame_standard, anchor="nw")
        canvas_standard.configure(yscrollcommand=scrollbar_standard.set)

        run_frame = ttk.LabelFrame(scrollable_frame_standard, text="800/1000米附加分条件（秒数）")
        run_frame.pack(fill="x", padx=20, pady=10)

        row = 0
        for label, default in RUN_BONUS.items():
            ttk.Label(run_frame, text=label + ":").grid(row=row, column=0, padx=10, pady=5, sticky="e")
            entry = ttk.Entry(run_frame, width=15)
            entry.insert(0, str(default))
            entry.grid(row=row, column=1, padx=10, pady=5)
            self.standard_entries[label] = entry
            row += 1

        other_frame = ttk.LabelFrame(scrollable_frame_standard, text="其他项目满分标准")
        other_frame.pack(fill="both", expand=True, padx=20, pady=10)

        row = 0
        for label, default in STANDARD_FULL.items():
            ttk.Label(other_frame, text=label + ":").grid(row=row, column=0, padx=10, pady=5, sticky="e")
            entry = ttk.Entry(other_frame, width=15)
            entry.insert(0, str(default))
            entry.grid(row=row, column=1, padx=10, pady=5)
            self.standard_entries[label] = entry
            row += 1

        # 单位配置
        unit_frame = ttk.LabelFrame(scrollable_frame_standard, text="单位配置")
        unit_frame.pack(fill="x", padx=20, pady=10)

        ttk.Label(unit_frame, text="立定跳远单位:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
        var_jump = tk.StringVar(value=DEFAULT_CONFIG['立定跳远单位'])
        ttk.Combobox(unit_frame, textvariable=var_jump, values=['厘米', '米'], state="readonly", width=10).grid(row=0, column=1, padx=10, pady=5)
        self.standard_entries['立定跳远单位'] = var_jump

        ttk.Label(unit_frame, text="实心球单位:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
        var_ball = tk.StringVar(value=DEFAULT_CONFIG['实心球单位'])
        ttk.Combobox(unit_frame, textvariable=var_ball, values=['米', '厘米'], state="readonly", width=10).grid(row=1, column=1, padx=10, pady=5)
        self.standard_entries['实心球单位'] = var_ball

       # 其他配置（新增更多项）
        extra_frame = ttk.LabelFrame(scrollable_frame_standard, text="其他配置")
        extra_frame.pack(fill="x", padx=20, pady=10)

        row = 0
        ttk.Label(extra_frame, text="满分标准:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        entry_full = ttk.Entry(extra_frame, width=15)
        entry_full.insert(0, str(DEFAULT_CONFIG['满分标准']))
        entry_full.grid(row=row, column=1, padx=10, pady=5)
        self.standard_entries['满分标准'] = entry_full
        row += 1

        ttk.Label(extra_frame, text="优秀标准:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        entry_excellent = ttk.Entry(extra_frame, width=15)
        entry_excellent.insert(0, str(DEFAULT_CONFIG['优秀标准']))
        entry_excellent.grid(row=row, column=1, padx=10, pady=5)
        self.standard_entries['优秀标准'] = entry_excellent
        row += 1

        ttk.Label(extra_frame, text="是否加排名列:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        var_rank = tk.BooleanVar(value=DEFAULT_CONFIG['是否加排名列'])
        ttk.Checkbutton(extra_frame, variable=var_rank).grid(row=row, column=1, padx=10, pady=5)
        self.standard_entries['是否加排名列'] = var_rank
        row += 1

        ttk.Label(extra_frame, text="缺考背景色（16进制）:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        entry_bg = ttk.Entry(extra_frame, width=15)
        entry_bg.insert(0, DEFAULT_CONFIG['缺考背景色'])
        entry_bg.grid(row=row, column=1, padx=10, pady=5)
        self.standard_entries['缺考背景色'] = entry_bg
        row += 1

        ttk.Label(extra_frame, text="缺考字体颜色（16进制）:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        entry_font = ttk.Entry(extra_frame, width=15)
        entry_font.insert(0, DEFAULT_CONFIG['缺考字体颜色'])
        entry_font.grid(row=row, column=1, padx=10, pady=5)
        self.standard_entries['缺考字体颜色'] = entry_font
        row += 1

        ttk.Label(extra_frame, text="总分小数位数:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        entry_decimal = ttk.Entry(extra_frame, width=15)
        entry_decimal.insert(0, str(DEFAULT_CONFIG['总分小数位数']))
        entry_decimal.grid(row=row, column=1, padx=10, pady=5)
        self.standard_entries['总分小数位数'] = entry_decimal
        row += 1

        ttk.Label(extra_frame, text="排名并列方式:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        var_rank_method = tk.StringVar(value=DEFAULT_CONFIG['排名并列方式'])
        ttk.Combobox(extra_frame, textvariable=var_rank_method, values=['min', 'dense'], state="readonly", width=10).grid(row=row, column=1, padx=10, pady=5)
        self.standard_entries['排名并列方式'] = var_rank_method
        row += 1

        ttk.Label(extra_frame, text="是否自动排序班级:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        var_sort = tk.BooleanVar(value=DEFAULT_CONFIG['是否自动排序班级'])
        ttk.Checkbutton(extra_frame, variable=var_sort).grid(row=row, column=1, padx=10, pady=5)
        self.standard_entries['是否自动排序班级'] = var_sort
        row += 1

        ttk.Label(extra_frame, text="统计包含优秀率:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        var_excellent_rate = tk.BooleanVar(value=DEFAULT_CONFIG['统计包含优秀率'])
        ttk.Checkbutton(extra_frame, variable=var_excellent_rate).grid(row=row, column=1, padx=10, pady=5)
        self.standard_entries['统计包含优秀率'] = var_excellent_rate
        row += 1

        ttk.Label(extra_frame, text="统计文件名后缀:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        entry_suffix = ttk.Entry(extra_frame, width=20)
        entry_suffix.insert(0, DEFAULT_CONFIG['统计文件名后缀'])
        entry_suffix.grid(row=row, column=1, padx=10, pady=5)
        self.standard_entries['统计文件名后缀'] = entry_suffix

        # 自定义项目（合并版：映射式+比例式，删除重复的旧版）
        custom_frame = ttk.LabelFrame(scrollable_frame_standard, text="自定义项目（支持映射式/比例式）")
        custom_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # 1. 映射式自定义项目（新增）
        map_frame = ttk.LabelFrame(custom_frame, text="映射式项目（一一对应评分）")
        map_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # 映射项目添加区域
        add_map_frame = ttk.Frame(map_frame)
        add_map_frame.pack(fill="x", pady=5)
        ttk.Label(add_map_frame, text="项目名:").grid(row=0, column=0, padx=5)
        self.map_name = ttk.Entry(add_map_frame, width=15)
        self.map_name.grid(row=0, column=1, padx=5)

        ttk.Label(add_map_frame, text="越大越好:").grid(row=0, column=2, padx=5)
        self.map_bigger = tk.BooleanVar(value=True)
        ttk.Checkbutton(add_map_frame, variable=self.map_bigger).grid(row=0, column=3, padx=5)

        ttk.Label(add_map_frame, text="映射规则（格式：阈值,得分;阈值,得分）:").grid(row=0, column=4, padx=5)
        self.map_rules = ttk.Entry(add_map_frame, width=50)
        self.map_rules.grid(row=0, column=5, padx=5)
        ttk.Button(add_map_frame, text="添加映射项目", command=self.add_map_project).grid(row=0, column=6, padx=10)

        # 映射项目列表
        self.map_listbox = tk.Listbox(map_frame, height=8)
        self.map_listbox.pack(fill="both", expand=True, padx=10, pady=5)
        ttk.Button(map_frame, text="删除选中映射项目", command=self.delete_map_project).pack(pady=5)

        # 2. 比例式自定义项目（保留原功能，兼容旧版）
        ratio_frame = ttk.LabelFrame(custom_frame, text="比例式项目（满分3分制）")
        ratio_frame.pack(fill="both", expand=True, padx=10, pady=5)

        add_ratio_frame = ttk.Frame(ratio_frame)
        add_ratio_frame.pack(fill="x", pady=5)
        ttk.Label(add_ratio_frame, text="项目名:").grid(row=0, column=0, padx=5)
        self.ratio_name = ttk.Entry(add_ratio_frame, width=15)
        self.ratio_name.grid(row=0, column=1, padx=5)

        ttk.Label(add_ratio_frame, text="满分值:").grid(row=0, column=2, padx=5)
        self.ratio_full = ttk.Entry(add_ratio_frame, width=10)
        self.ratio_full.grid(row=0, column=3, padx=5)

        ttk.Label(add_ratio_frame, text="单位:").grid(row=0, column=4, padx=5)
        self.ratio_unit = ttk.Entry(add_ratio_frame, width=10)
        self.ratio_unit.grid(row=0, column=5, padx=5)

        ttk.Label(add_ratio_frame, text="越大越好:").grid(row=0, column=6, padx=5)
        self.ratio_bigger = tk.BooleanVar(value=True)
        ttk.Checkbutton(add_ratio_frame, variable=self.ratio_bigger).grid(row=0, column=7, padx=5)

        ttk.Button(add_ratio_frame, text="添加比例项目", command=self.add_ratio_project).grid(row=0, column=8, padx=10)

        # 比例项目列表
        self.ratio_listbox = tk.Listbox(ratio_frame, height=8)
        self.ratio_listbox.pack(fill="both", expand=True, padx=10, pady=5)
        ttk.Button(ratio_frame, text="删除选中比例项目", command=self.delete_ratio_project).pack(pady=5)

        canvas_standard.pack(side="left", fill="both", expand=True)
        scrollbar_standard.pack(side="right", fill="y")

    def add_custom_project(self):
        name = self.custom_name.get().strip()
        if not name:
            messagebox.showwarning("警告", "请输入项目名")
            return
        try:
            full = float(self.custom_full.get().strip())
        except:
            messagebox.showwarning("警告", "满分值必须是数字")
            return
        unit = self.custom_unit.get().strip() or "个"
        bigger = self.custom_bigger.get()

        project = {"name": name, "full_value": full, "unit": unit, "bigger_better": bigger}
        self.custom_projects.append(project)
        self.update_custom_list()
        self.custom_name.delete(0, tk.END)
        self.custom_full.delete(0, tk.END)
        self.custom_unit.delete(0, tk.END)

    def delete_custom_project(self):
        selected = self.custom_listbox.curselection()
        if selected:
            del self.custom_projects[selected[0]]
            self.update_custom_list()

    def update_custom_list(self):
        self.custom_listbox.delete(0, tk.END)
        for proj in self.custom_projects:
            direction = "越大越好" if proj['bigger_better'] else "越小越好"
            text = f"{proj['name']} | 满分 {proj['full_value']} {proj['unit']} | {direction}"
            self.custom_listbox.insert(tk.END, text)

    def toggle_stats_path(self):
        if self.var_statistics.get():
            self.stats_entry.config(state="normal")
            if self.input_file:
                dir_name = os.path.dirname(self.input_file)
                base_name = os.path.basename(self.input_file).rsplit('.', 1)[0]
                suggested = os.path.join(dir_name, f"{base_name}_统计汇总.xlsx")
                self.stats_file = suggested
                self.stats_entry.delete(0, tk.END)
                self.stats_entry.insert(0, suggested)
        else:
            self.stats_entry.config(state="disabled")
            self.stats_file = ""

    def select_stats_output(self):
        if self.var_statistics.get():
            file = filedialog.asksaveasfilename(title="保存统计汇总", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if file:
                self.stats_file = file
                self.stats_entry.delete(0, tk.END)
                self.stats_entry.insert(0, file)

    def save_config(self):
        config = {}
        config['当前版本'] = DEFAULT_CONFIG['当前版本']
        config['忽略更新版本'] = DEFAULT_CONFIG.get('忽略更新版本', '')
        # 列配置
        for label, entry in self.config_entries.items():
            value = entry.get().strip().upper()
            if value:
                config[label] = value
        # 标准配置
        for label, entry in self.standard_entries.items():
            if isinstance(entry, (tk.StringVar, tk.BooleanVar)):
                config[label] = entry.get()
            else:
                value = entry.get().strip()
                if value:
                    try:
                        config[label] = float(value)
                    except:
                        config[label] = value
        # 自定义项目（比例式+映射式）
        config['自定义项目'] = DEFAULT_CONFIG['自定义项目']
        config['自定义映射项目'] = DEFAULT_CONFIG['自定义映射项目']
        # 保存
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=4)
        messagebox.showinfo("成功", "配置已保存！")

    def load_config(self):
        if not os.path.exists(CONFIG_FILE):
            self.update_map_list()
            self.update_ratio_list()
            return
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                saved = json.load(f)
            # 列配置
            for label, entry in self.config_entries.items():
                if label in saved:
                    entry.delete(0, tk.END)
                    entry.insert(0, saved[label])
            # 标准配置
            for label, entry in self.standard_entries.items():
                if label in saved:
                    if isinstance(entry, (tk.StringVar, tk.BooleanVar)):
                        entry.set(saved[label])
                    else:
                        entry.delete(0, tk.END)
                        entry.insert(0, str(saved[label]))
            # 自定义项目
            DEFAULT_CONFIG['自定义项目'] = saved.get('自定义项目', [])
            DEFAULT_CONFIG['自定义映射项目'] = saved.get('自定义映射项目', [])
            self.update_map_list()
            self.update_ratio_list()
        except Exception as e:
            messagebox.showerror("错误", f"加载配置失败：{e}")

    def select_input(self):
        file = filedialog.askopenfilename(title="选择原始Excel文件", filetypes=[("Excel files", "*.xlsx *.xls")])
        if file:
            self.input_file = file
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, file)
            self.load_sheets()

            dir_name = os.path.dirname(file)
            base_name = os.path.basename(file).rsplit('.', 1)[0]
            suggested = os.path.join(dir_name, f"{base_name}_已评分.xlsx")
            self.output_file = suggested
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, suggested)

            if self.var_statistics.get():
                stats_suggested = os.path.join(dir_name, f"{base_name}_统计汇总.xlsx")
                self.stats_file = stats_suggested
                self.stats_entry.delete(0, tk.END)
                self.stats_entry.insert(0, stats_suggested)

    def load_sheets(self):
        try:
            xl = pd.ExcelFile(self.input_file)
            self.sheets = xl.sheet_names
            for widget in self.sheet_frame.winfo_children():
                widget.destroy()
            self.sheet_vars = []
            for sheet in self.sheets:
                var = tk.IntVar(value=1)
                cb = ttk.Checkbutton(self.sheet_frame, text=sheet, variable=var)
                cb.pack(anchor="w")
                self.sheet_vars.append(var)
        except Exception as e:
            messagebox.showerror("错误", f"读取文件失败：{e}")

    def select_output(self):
        file = filedialog.asksaveasfilename(title="保存评分结果", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.output_file = file
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, file)

    def get_current_config(self):
        config = {}
        for label, entry in self.config_entries.items():
            value = entry.get().strip().upper()
            if value:
                config[label] = value

        name_col = config.get('姓名列', 'C')
        gender_col = config.get('性别列', 'D')
        proj_cols = [config.get('第一类项目列', 'E'), config.get('第二类项目列', 'I'),
                     config.get('第三类项目列', 'M'), config.get('第四类项目列', 'Q')]
        res_cols = [config.get('第一类成绩列', 'F'), config.get('第二类成绩列', 'J'),
                    config.get('第三类成绩列', 'N'), config.get('第四类成绩列', 'R')]
        score_cols = [config.get('第一类得分列', 'G'), config.get('第二类得分列', 'K'),
                      config.get('第三类得分列', 'O'), config.get('第四类得分列', 'S')]
        total_col = config.get('总分列', 'T')

        # 更新全局配置
        DEFAULT_CONFIG['立定跳远单位'] = self.standard_entries['立定跳远单位'].get()
        DEFAULT_CONFIG['实心球单位'] = self.standard_entries['实心球单位'].get()
        DEFAULT_CONFIG['自定义项目'] = self.custom_projects

        try:
            DEFAULT_CONFIG['满分标准'] = float(self.standard_entries['满分标准'].get())
            DEFAULT_CONFIG['优秀标准'] = float(self.standard_entries['优秀标准'].get())
            DEFAULT_CONFIG['是否加排名列'] = self.standard_entries['是否加排名列'].get()
            DEFAULT_CONFIG['缺考背景色'] = self.standard_entries['缺考背景色'].get().upper()
        except:
            pass

        # 更新附加分标准
        global RUN_BONUS, STANDARD_FULL
        for label, entry in self.standard_entries.items():
            if label in RUN_BONUS or label in STANDARD_FULL:
                try:
                    RUN_BONUS[label] = float(entry.get())
                except:
                    pass

        return (name_col, gender_col, proj_cols, res_cols, score_cols, total_col)

    def process(self):
        if not self.input_file:
            messagebox.showwarning("警告", "请先选择输入文件！")
            return
        if not self.output_file:
            messagebox.showwarning("警告", "请先选择输出文件！")
            return

        selected_indices = [i for i, var in enumerate(self.sheet_vars) if var.get() == 1]
        if not selected_indices:
            messagebox.showwarning("警告", "请至少选择一个班级！")
            return

        selected_sheets = [self.sheets[i] for i in selected_indices]

        NAME_COL, GENDER_COL, PROJ_COLS, RES_COLS, SCORE_COLS, TOTAL_COL = self.get_current_config()

        NAME_N = col_to_num(NAME_COL)
        GENDER_N = col_to_num(GENDER_COL)
        PROJ_N = [col_to_num(c) for c in PROJ_COLS]
        RES_N = [col_to_num(c) for c in RES_COLS]
        SCORE_N = [col_to_num(c) for c in SCORE_COLS]
        TOTAL_N = col_to_num(TOTAL_COL)

        # 从配置读取更多自定义项
        decimal_places = int(DEFAULT_CONFIG.get('总分小数位数', 2))
        rank_method = DEFAULT_CONFIG.get('排名并列方式', 'min')
        auto_sort = DEFAULT_CONFIG.get('是否自动排序班级', True)
        excellent_rate_in_stats = DEFAULT_CONFIG.get('统计包含优秀率', True)
        stats_suffix = DEFAULT_CONFIG.get('统计文件名后缀', '_统计汇总')

        messagebox.showinfo("开始", f"正在处理 {len(selected_sheets)} 个班级，请稍等...")

        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            for sheet_name in selected_sheets:
                df = pd.read_excel(self.input_file, sheet_name=sheet_name, header=None, skiprows=1, dtype=str)

                max_col = max(TOTAL_N, max(SCORE_N + RES_N + PROJ_N, default=0))
                while len(df.columns) <= max_col:
                    df[len(df.columns)] = np.nan

                # 修复点1：新增列标记是否缺考，避免行号错乱
                df['是否缺考'] = False

                for i in range(len(df)):
                    if pd.isna(df.iat[i, NAME_N]): 
                        continue
                    gender = str(df.iat[i, GENDER_N]).strip()
                    if gender not in ['男', '女']: 
                        continue

                    scores_num = []
                    has_missing = False

                    for k in range(4):
                        proj = df.iat[i, PROJ_N[k]]
                        res = df.iat[i, RES_N[k]]
                        sc = get_score(gender, proj, res)
                        df.iat[i, SCORE_N[k]] = sc
                        if sc == "缺考":
                            scores_num.append(0.0)
                            has_missing = True
                        else:
                            scores_num.append(sc if isinstance(sc, (int, float)) else 0.0)

                    total = round(sum(scores_num), decimal_places)
                    df.iat[i, TOTAL_N] = total
                    # 修复点2：用列标记缺考，而非记录行号
                    if has_missing:
                        df.iat[i, df.columns.get_loc('是否缺考')] = True

                # 先设置列名（关键！避免 KeyError: '总分'）
                df.columns = HEADERS[:len(df.columns)-1] + ['是否缺考']  # 兼容新增列

                # 再加排名列（可选 + 自定义方式 + 自动排序）
                if DEFAULT_CONFIG['是否加排名列']:
                    df['总分'] = pd.to_numeric(df['总分'], errors='coerce')
                    df['排名'] = df['总分'].rank(method=rank_method, ascending=False)
                    if auto_sort:
                        df = df.sort_values('总分', ascending=False).reset_index(drop=True)
                    df['排名'] = df['排名'].astype(int)

                # 修复点3：导出时排除标记列，避免污染Excel
                df_export = df.drop(columns=['是否缺考'], errors='ignore')
                df_export.to_excel(writer, sheet_name=sheet_name, index=False)

        # 标红标黄（修复核心：直接读取导出后的Excel，按得分列是否缺考判断）
        wb = load_workbook(self.output_file)
        red_font = Font(color=DEFAULT_CONFIG['缺考字体颜色'])
        fill = PatternFill("solid", fgColor=DEFAULT_CONFIG['缺考背景色'])
        
        for sheet_name in selected_sheets:
            ws = wb[sheet_name]
            # 先获取所有缺考的行号（从Excel直接判断，避免行号错乱）
            absent_rows = set()
            # 遍历得分列，收集有缺考的行号
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    col_idx = cell.column - 1
                    if col_idx in SCORE_N and cell.value == "缺考":
                        absent_rows.add(cell.row)
                        cell.font = red_font
                        cell.fill = fill
            
            # 给有缺考的行的总分列标黄
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    col_idx = cell.column - 1
                    if col_idx == TOTAL_N and cell.row in absent_rows:
                        cell.font = red_font
                        cell.fill = fill

            # 自动列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 40)

        wb.save(self.output_file)

        # 班级统计汇总（单独文件，可选优秀率）
        if self.var_statistics.get() and self.stats_file:
            stats = []
            full_threshold = DEFAULT_CONFIG['满分标准']
            excellent_threshold = DEFAULT_CONFIG['优秀标准']
            for sheet_name in selected_sheets:
                df = pd.read_excel(self.output_file, sheet_name=sheet_name)
                total_students = len(df)
                valid_total = pd.to_numeric(df['总分'], errors='coerce')
                avg = valid_total.mean()
                full = (valid_total >= full_threshold).sum()
                excellent = (valid_total >= excellent_threshold).sum()
                missing = (df.iloc[:, [6,10,14,18]] == "缺考").sum().sum()
                max_score = valid_total.max()
                min_score = valid_total.min()

                row_data = [sheet_name, total_students, round(avg, 2), full, excellent, missing, max_score, min_score]
                if excellent_rate_in_stats and total_students > 0:
                    excellent_rate_pct = round(excellent / total_students * 100, 2)
                    row_data.append(excellent_rate_pct)

                stats.append(row_data)

            columns = ['班级', '人数', '平均分', '满分人数', '优秀人数', '缺考项目数', '最高分', '最低分']
            if excellent_rate_in_stats:
                columns.append('优秀率(%)')

            stats_df = pd.DataFrame(stats, columns=columns)
            stats_df.to_excel(self.stats_file, index=False)
            os.startfile(self.stats_file)

        messagebox.showinfo("成功", "所有任务完成！")
        os.startfile(self.output_file)

    # 映射式项目操作
    def add_map_project(self):
        name = self.map_name.get().strip()
        rules = self.map_rules.get().strip()
        if not name or not rules:
            messagebox.showwarning("警告", "项目名和映射规则不能为空")
            return
        try:
            # 解析映射规则（格式：阈值,得分;阈值,得分）
            score_map = []
            for rule in rules.split(';'):
                lim, sc = rule.strip().split(',')
                score_map.append((float(lim), float(sc)))
            # 按阈值排序（越大越好升序，越小越好降序）
            bigger = self.map_bigger.get()
            score_map.sort(key=lambda x: x[0], reverse=not bigger)
            # 添加到配置
            project = {
                'name': name,
                'score_map': score_map,
                'bigger_better': bigger
            }
            DEFAULT_CONFIG['自定义映射项目'].append(project)
            self.update_map_list()
            # 清空输入
            self.map_name.delete(0, tk.END)
            self.map_rules.delete(0, tk.END)
        except Exception as e:
            messagebox.showerror("错误", f"映射规则格式错误（示例：7.1,3.0;7.2,2.85）：{e}")

    def delete_map_project(self):
        selected = self.map_listbox.curselection()
        if selected:
            del DEFAULT_CONFIG['自定义映射项目'][selected[0]]
            self.update_map_list()

    def update_map_list(self):
        self.map_listbox.delete(0, tk.END)
        for proj in DEFAULT_CONFIG['自定义映射项目']:
            direction = "越大越好" if proj['bigger_better'] else "越小越好"
            rules = " | ".join([f"{lim}:{sc}分" for lim, sc in proj['score_map']])
            text = f"{proj['name']} | {direction} | 规则：{rules}"
            self.map_listbox.insert(tk.END, text)

    # 比例式项目操作（替换原add_custom_project等方法）
    def add_ratio_project(self):
        name = self.ratio_name.get().strip()
        if not name:
            messagebox.showwarning("警告", "请输入项目名")
            return
        try:
            full = float(self.ratio_full.get().strip())
        except:
            messagebox.showwarning("警告", "满分值必须是数字")
            return
        unit = self.ratio_unit.get().strip() or "个"
        bigger = self.ratio_bigger.get()
        project = {"name": name, "full_value": full, "unit": unit, "bigger_better": bigger}
        DEFAULT_CONFIG['自定义项目'].append(project)
        self.update_ratio_list()
        self.ratio_name.delete(0, tk.END)
        self.ratio_full.delete(0, tk.END)
        self.ratio_unit.delete(0, tk.END)

    def delete_ratio_project(self):
        selected = self.ratio_listbox.curselection()
        if selected:
            del DEFAULT_CONFIG['自定义项目'][selected[0]]
            self.update_ratio_list()

    def update_ratio_list(self):
        self.ratio_listbox.delete(0, tk.END)
        for proj in DEFAULT_CONFIG['自定义项目']:
            direction = "越大越好" if proj['bigger_better'] else "越小越好"
            text = f"{proj['name']} | 满分 {proj['full_value']} {proj['unit']} | {direction}"
            self.ratio_listbox.insert(tk.END, text)

    def check_for_update(self):
        current_version = DEFAULT_CONFIG['当前版本']
        ignore_version = DEFAULT_CONFIG.get('忽略更新版本', '')

        try:
            response = requests.get("https://1427.tech/projects/PEScoring/latest", timeout=5)
            if response.status_code == 200:
                latest_version = response.text.strip()
                if latest_version != current_version and latest_version != ignore_version:
                    result = messagebox.askyesnocancel(
                        "发现新版本",
                        f"当前版本：{current_version}\n最新版本：{latest_version}\n\n是否打开下载页面？"
                    )
                    if result is True:  # 是 → 打开浏览器
                        import webbrowser
                        webbrowser.open("https://1427.tech/projects/PEScoring")
                    elif result is False:  # 否 → 不再提醒此版本
                        DEFAULT_CONFIG['忽略更新版本'] = latest_version
                        self.save_config()  # 保存到配置文件
        except:
            pass  # 网络错误或超时，静默忽略

if __name__ == "__main__":
    root = tk.Tk()
    app = SportsScoreGUI(root)
    root.mainloop()





